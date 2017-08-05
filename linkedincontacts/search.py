import bs4
from selenium import webdriver
import time
from selenium.webdriver.common.keys import Keys
from selenium.common.exceptions import NoSuchElementException
import openpyxl
import random
import logging

random.seed()

username = 'Username@xxx.com'
password = 'passowrd'
URL = 'https://www.linkedin.com/uas/login'
dstnfile = 'search.xlsx'
baseurl = 'https://www.linkedin.com'
logfilename = 'script.log'
logging.basicConfig(filename=logfilename, level=logging.INFO, filemode='w', format='%(asctime)s - %(levelname)s - %('
                                                                                   'message)s')
logger = logging.getLogger()

# Login into linkedin using selenium
driver = webdriver.Firefox()
driver.get(URL)
elem = driver.find_element_by_id('session_key-login')
elem.send_keys(username)
elem = driver.find_element_by_id('session_password-login')
elem.send_keys(password)
elem.send_keys(Keys.RETURN)
time.sleep(5)

input('Navigate to the search page for people & then press enter : ')

# Setup Workbook

# Write data to xlsx file
writebook = openpyxl.Workbook()
writesheet = writebook.active
writesheet.cell(row=1, column=1, value='Name')
writesheet.cell(row=1, column=2, value='Profile Link')
writesheet.cell(row=1, column=3, value='Title')

iterator = 2
flag = True

while iterator <= 101:
    driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
    time.sleep(2)
    html = driver.page_source
    soup = bs4.BeautifulSoup(html, 'html.parser')
    allentities = soup.find_all('div', class_='search-result__info')
    for entity in allentities:
        name = entity.select('span.name.actor-name')[0].text
        namelink = baseurl + entity.find('a', class_='search-result__result-link')['href']
        headlinelist = entity.find('p', class_='subline-level-1').contents
        headline = ' '.join([bs4.BeautifulSoup(str(x), 'html.parser').text for x in headlinelist])

        # Write Data
        writesheet.cell(row=iterator, column=1, value=name)
        writesheet.cell(row=iterator, column=2, value=namelink)
        writesheet.cell(row=iterator, column=3, value=headline)
        logger.info(name)
        logger.info(namelink)
        logger.info(headlinelist)
        logger.info(headline)

        # Update
        iterator += 1
    try:
        elem = driver.find_element_by_css_selector('button.next')
        elem.click()
        logger.info('iterator = ' + str(iterator))
        time.sleep(random.randint(15, 30))
    except NoSuchElementException:
        logger.warning('Found exception End of results')
        break

writebook.save(dstnfile)

# try:
#     elem = WebDriverWait(driver, 10).until(
#         EC.presence_of_element_located((By.CSS_SELECTOR, 'form.nav-search'))
#     )
# except Exception as e:
#     print (str(e))
#     print("Unable to find navigation bar for searching after 30 sec of wait, Exiting .... ")
#     exit()
#
# print (type(elem))
# searchelem = elem.find_element_by_tag_name('input')
# print(searchelem.tag_name + searchelem.get_attribute('id'))
# searchelem.send_keys('CEO')
# searchelem.send_keys(Keys.RETURN)
# html = driver.page_source
# print (driver.current_url)

# from bs4 import BeautifulSoup
# from selenium import webdriver
# driver = webdriver.Firefox() #I actually used the chromedriver and did not test firefox, but it should work.
# profile_link="https://www.linkedin.com/in/johnsmith1"
# driver.get(profile_link)
# html=driver.page_source
# soup=BeautifulSoup(html) #specify parser or it will auto-select for you
# summary=soup.find('section', { "id" : "summary" })
# print summary.getText()



# cookie_filename = 'parser.cookies.txt'


# class LinkedInParser:
#     def loadSoup(self, url, data=None):
#         """
#         Combine loading of URL, HTML, and parsing with BeautifulSoup
#         """
#         html = self.loadPage(url, data)
#         soup = bs4.BeautifulSoup(html, "html.parser")
#         return soup
#
#     def loadPage(self, url, data=None):
#         """
#         Utility function to load HTML from URLs for us with hack to continue despite 404
#         """
#         # We'll print the url in case of infinite loop
#         # print "Loading URL: %s" % url
#         try:
#             if data is not None:
#                 response = self.opener.open(url, data)
#             else:
#                 response = self.opener.open(url)
#             return ''.join([str(l) for l in response.readlines()])
#         except Exception as e:
#             # If URL doesn't load for ANY reason, try again...
#             # Quick and dirty solution for 404 returns because of network problems
#             # However, this could infinite loop if there's an actual problem
#             return self.loadPage(url, data)
#
#     def loginPage(self):
#         """
#         Handle login. This should populate our cookie jar
#         :param self:
#         :return:
#         """
#         soup = self.loadSoup("https://www.linkedin.com/")
#         csrf = soup.find(id="loginCsrfParam-login")['value']
#         login_data = urllib.parse.urlencode({
#             'session_key': self.login,
#             'session_password': self.password,
#             'loginCsrfParam': csrf,
#         }).encode('utf8')
#
#         self.loadPage("https://www.linkedin.com/uas/login-submit", login_data)
#         return
#
#     def loadTitle(self):
#         soup = self.loadSoup("https://www.linkedin.com/feed/")
#         f = open('linkedin.html', 'w')
#         f.write(soup.prettify())
#         f.close()
#         return soup.find("title")
#
#     def __init__(self, login, password):
#         """Start up..."""
#         self.login = login
#         self.password = password
#         # Simulate browser with cookies enabled
#         self.cj = http.cookiejar.MozillaCookieJar(cookie_filename)
#         if os.access(cookie_filename, os.F_OK):
#             self.cj.load()
#         self.opener = urllib.request.build_opener(
#             urllib.request.HTTPRedirectHandler(),
#             urllib.request.HTTPHandler(debuglevel=0),
#             urllib.request.HTTPSHandler(debuglevel=0),
#             urllib.request.HTTPCookieProcessor(self.cj)
#         )
#         self.opener.addheaders = [
#             ('User-Agent', 'Mozilla/5.0 (X11; Linux x86_64; rv:52.0) Gecko/20100101 Firefox/52.0')
#         ]
#         # Login
#         self.loginPage()
#         title = self.loadTitle()
#         print(title)
#         self.cj.save()

# session = requests.Session()
# headers = {'User-Agent': 'Mozilla/5.0 (X11; Linux x86_64; rv:52.0) Gecko/20100101 Firefox/52.0'}
#
# loginresponse = session.get(URL, headers=headers)
# soup = bs4.BeautifulSoup(loginresponse.text, 'html.parser')
# f = open('linkedin1.html', 'w')
# f.write(soup.prettify())
# f.close()
#
# # Get hidden form inputs
# inputs = soup.find('form', {'name': 'login'}).find_all('input', {'type': ['hidden', 'submit']})
#
# # Create POST data
# post = {inp.get('name'): inp.get('value') for inp in inputs}
# post['session_key'] = username
# post['session_password'] = password
# # Post login
# post_response = session.post('https://www.linkedin.com/uas/login-submit', data=post, headers=headers)
# if post_response.status_code == requests.codes.ok:
#     print("All is well")
#     print(post_response.status_code)
# else:
#     print("Failed , not again")
# # post_response = session.post('https://www.linkedin.com/uas/login-submit', data=post)
#
# # Get home page
# home_response = session.get('https://linkedin.com/nhome/')
# home = bs4.BeautifulSoup(home_response.text, 'html.parser')
